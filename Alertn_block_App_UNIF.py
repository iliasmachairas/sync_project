# -*- coding: utf-8 -*-
"""
Created on Fri Jun 10 20:37:08 2022

@author: ilias
"""
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import datetime
#from io import BytesIO
import numpy as np
import geopandas as gpd
import matplotlib.pyplot as plt
import matplotlib
from PIL import Image

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# draft comment 

st.set_page_config(layout = "wide")

st.markdown("""Alternating Block method - Greece""")

col2, space2, col3, space3, col4  = st.columns((8,1,10,1,10))

stations = pd.read_excel('GR_StationsIDF_FD_English.xlsx')
stations_id = stations.loc[:,'id'].values.tolist()

stations_shp = gpd.read_file('GR_StationsIDF_FD.shp')
stations_shp_4326 = stations_shp.to_crs(epsg='4326')

with col2:
    
    #st.markdown("""Select rainfall station""", selected_station)
    selected_id = st.selectbox(label='Select station id', options=stations_id)
    selected_station = stations.loc[stations['id'] == selected_id]
   
    # select return period
    min_T, mean_T, max_T = 5, 50, 100
    selected_T = st.slider("Select Return Period (T)", min_value=min_T,
    max_value=max_T, value=mean_T, step=5)
    
    # select time step
    min_dur, mean_dur, max_dur = 5, 10, 20
    sel_time_step = st.slider("Select time step (minutes)", min_value=min_dur,
    max_value=max_dur, value=mean_dur, step=5)
    
     # select storm duration
    storm_dur_options = np.arange(1,37,1)
    selected_storm_dur = st.selectbox(label='Select storm duration (h)',
                                options=storm_dur_options, index=5)
    
    # M ap creation
    
    
    # # selection column
    stations_shp_4326['selection'] = 'No'
    stations_shp_4326.loc[stations_shp_4326['Id'] == int(selected_id), 'selection'] = 'yes'
    
    # # Size column
    stations_shp_4326['size'] = 4
    stations_shp_4326.loc[stations_shp_4326['Id'] == int(selected_id), 'size'] = 16
    
    # convert wgs-84
    stations_shp_4326['x_wgs84'] = stations_shp_4326.geometry.x
    stations_shp_4326['y_wgs84'] = stations_shp_4326.geometry.y
    
    center = stations_shp_4326.loc[stations_shp_4326['Id'] ==  int(selected_id)]
    center_dict = {'lon':center.x_wgs84.values[0], 'lat':center.y_wgs84.values[0]}
    
    # Parameters
    name = selected_station['name'].values[0]
    kappa = selected_station['kappa'].values[0]
    lambda_value = selected_station['lambda'].values[0]
    psi = selected_station['psi'].values[0]
    theta = selected_station['theta'].values[0]
    itta = selected_station['itta'].values[0]

    st.markdown('Parameters of the station')
    st.write('name=',name)
    st.write('kappa=',np.round(kappa,2),'\t', 'lambda=', np.round(lambda_value,2),
             'psi=',np.round(psi,2))
    st.write('theta=',np.round(theta,2),'itta=',np.round(itta,2))
    

    
with col3:
    
    
    def rain_intens(T,d):
        a = lambda_value * (T**kappa - psi)
        b = (1 + d / theta) ** itta
        return a / b

    rain_intens_v = np.vectorize(rain_intens)
    d = np.linspace(0,24,100)
    


    
    # IDF chart
    fig = go.Figure(layout=go.Layout(
        title=go.layout.Title(text="IDF Curves - "+str(selected_station['name'].values.tolist()[0]))))

    y = rain_intens_v(selected_T,d)
    fig.add_scatter(x=d, y=y, name='T='+str(selected_T))

    fig.update_layout(
        xaxis_title="Duration (h)",
        yaxis_title="Rainfall Intensity (mm/h)",
        legend_title="Return Periods",
        width=400, height=400)
    
    st.plotly_chart(fig, use_container_width=True)
    
    st.write('Station Map')
    fig_map = px.scatter_mapbox(stations_shp_4326, lon='x_wgs84', lat='y_wgs84',
                    color='selection', hover_name='Id', hover_data=["Name"], size='size',
                            center=center_dict, zoom=7, height=500)

    fig_map.update_layout(mapbox_style="open-street-map",
                          width=600, height=350)
    st.plotly_chart(fig_map, use_container_width=True)
    
    # Alternating block method
    # Pre-processing
    timestep_h = sel_time_step / 60
    # t is an array of the time steps needed in hours
    t = np.linspace(timestep_h, selected_storm_dur, int(selected_storm_dur/timestep_h))
    rain_i = rain_intens_v(selected_T, t) * t # rain intensity
    
    storm_duration_min = selected_storm_dur*60
    tot_timesteps_min = storm_duration_min / sel_time_step
    timesteps_min = np.linspace(sel_time_step, storm_duration_min, int(tot_timesteps_min))
    
    rain_intens_1st_step = np.array([rain_i[0]])
    # I also kept the first value besides teh differences
    differences = np.diff(rain_i)
    rain_intens_steps = np.concatenate((rain_intens_1st_step,differences))
    merge_array = np.column_stack((rain_intens_steps, timesteps_min))
    rain_steps_df = pd.DataFrame(data = merge_array, index=t,
                             columns=['precip', 'steps'])
    
    
    # Processing - Alternative block method
    atltern_array = np.zeros(rain_steps_df.shape[0]) 
    rain_steps_df_sorted = rain_steps_df.sort_values(by='precip', ascending=False)
    mean_index = int(rain_steps_df.shape[0] /2)
    
    switch = True
    l1 = 1 
    l2 = 1
    for i in range(rain_steps_df_sorted.shape[0]):
        if  i == 0:
            atltern_array[mean_index] = rain_steps_df_sorted['precip'].iloc[0]
        else:
            if switch == True:
                atltern_array[mean_index - l1] = rain_steps_df_sorted['precip'].iloc[i]
                switch = False
                l1+=1
            else:
                atltern_array[mean_index + l2] = rain_steps_df_sorted['precip'].iloc[i]
                switch = True
                l2+=1            
    
    #atltern_df = pd.DataFrame(atltern_array, index=np.arange(1,rain_steps_df_sorted.shape[0]+1,1), columns= ['precip'])
    atltern_ar_merge = np.column_stack((atltern_array, timesteps_min))
    atltern_df = pd.DataFrame(atltern_ar_merge, index=t, columns= ['precip (mm)','time_step (min)'])
    cols = ['time_step (min)','precip (mm)',]
    atltern_df = atltern_df[cols]
    atltern_excel = atltern_df.to_excel('Alternating_block_method_app.xlsx',
                                        float_format='%.2f',
                                       index_label='time step (h)')
    
    # Export to excel
    #columns=['col1', 'col3']
    
    
with col4:
    
    # Show plot
    fig_Altern = go.Figure(layout=go.Layout(
        title=go.layout.Title(text="Alternating block method ")))
    
    fig_Altern.add_bar(x=atltern_df.index, y=atltern_df['precip (mm)'],
                           name='T='+str(selected_T))

    fig_Altern.update_layout(
        xaxis_title="Duration (h)",
        yaxis_title="Rainfall Dh (mm)",
        width=400, height=350)
    
    st.plotly_chart(fig_Altern, use_container_width=True)
    
    # Image plot
    matplotlib.rcParams.update({'font.size': 16})
    
    plt.figure(figsize=(14,10))
    plt.bar(atltern_df.index, atltern_df['precip (mm)'], width=0.2)
    plt.xlabel('Duration (h)')
    plt.ylabel('Rainfall (mm)')
    plt.title('Hyetograph')
    plt.savefig('Hydrograph_barplot.png')
    
    # -----  It was not necesssary to use it  ----
    #img = Image.open('Hydrograph_barplot.png')
    # from io import BytesIO
    # buf = BytesIO()
    # img.save(buf, format="JPEG")
    # byte_im = buf.getvalue()
    
    # --- Image with IDF_curves - 4 periods
    periods = [10, 25, 50, 100]
    d2 = np.linspace(0,16,100)
    plt.figure(figsize=(14,10))
    for i in periods:
        y_i = rain_intens_v(i,d2)
        plt.plot(d2, y_i, label='T'+str(i))
    plt.xlabel('Duration (h)')
    plt.ylabel('Rainfall intensity (mm/h)')
    plt.title('IDF Curves - Station: '+str(name))
    plt.legend()
    plt.savefig('IDF_curves.png')
    
    
    # Excel output configuration
    wb = load_workbook('Alternating_block_method_app.xlsx')
    ws = wb['Sheet1']
    ws.title = 'Alternating_block'
    ws = wb['Alternating_block']
    wb.create_sheet('Plot')
    ws['D3'] = 'Station'
    ws['D3'].font = Font(name="Arial", size=14, color="00FF0000")
    ws['D3'].alignment = Alignment(horizontal='center')
    ws['D4'] = str(selected_station['name'].values[0])
    
    # applying border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    max_row = wb['Alternating_block'].max_row
    for j in range(1,4):
        for i in range(1,max_row+1):
            ws.cell(column=j, row=i).border = thin_border
            if j==3:
                ws.cell(column=j, row=i).alignment = Alignment(horizontal='right')
            else:
                 ws.cell(column=j, row=i).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D4:E4')
    ws.merge_cells('D5:E5')
    ws['D4'].alignment = Alignment(horizontal='center')
    
    # Descriptive statitsics
    ws['G2'] = 'Descriptive Statistics'
    ws['G2'].alignment = Alignment(horizontal='center')
    ws['G2'].font = Font(bold=True)
    names = ['min', 'max', 'mean', 'std']
    
    column = 7
    for i, value in enumerate(names):
        cell = ws.cell(column=column, row=3+i, value=value)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='16A085', end_color='16A085', fill_type="solid")
        cell.font = Font(bold=True)
        
    # Descriptive stats
    ws['H3'].value = round(atltern_df['precip (mm)'].min(),2)
    ws['H4'].value = round(atltern_df['precip (mm)'].max(),2)
    ws['H5'].value = round(atltern_df['precip (mm)'].mean(),2)
    ws['H6'].value = round(np.std(atltern_df['precip (mm)']),2)
    
    for row in range(3, 7):
        ws["H{}".format(row)].number_format = '#,##0.00'
    
    # Number format
    for row in range(1, max_row+1):
        ws["A{}".format(row)].number_format = '#,##0.00'
        ws["C{}".format(row)].number_format = '#,##0.00'
    
    # Adjusting column width size
    column_widths = []
    for row in wb['Alternating_block'].iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    
    for i, column_width in enumerate(column_widths):
        wb['Alternating_block'].column_dimensions[get_column_letter(i + 1)].width = column_width  + 5 
    
    

    
# Creating a barchart
    chart1 = BarChart()
    data = Reference(ws, min_col = 3, min_row = 2, max_row = max_row) # You need to include the name of the column as well
    # besides the data
    cats = Reference(ws, min_col = 1, min_row = 3, max_row = max_row)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.title = "Alternating_Block_Method - Hyetograph"
    chart1.x_axis.title = 'Time (h)'
    chart1.y_axis.title = 'Rainfall (mm)'
    wb['Plot'].add_chart(chart1, "C4")
    
    openpyxl.chart.legend.Legend(legendEntry=())
    
    
    wb['Alternating_block'].insert_rows(1)
    # Freeze header
    wb["Alternating_block"].freeze_panes = "C3"
    # define printing area
    #wb['Alternating_block'].print_area = "A1:I27"
    
    #wb['Alternating_block'].insert_cols(1)
    
    # Fix the error with the plot 
    
    wb.save('Alternating_block_method_app.xlsx')
      
    # Download excel file     
    with open('Alternating_block_method_app.xlsx', 'rb') as f:
        st.download_button('Download Excel file', f, file_name = 'Altern_block.xlsx',
                           mime = 'application/vnd.ms-excel')
        
    btn_hydrograph = st.download_button(
      label="Download Hyetograph Image",
      data=open('Hydrograph_barplot.png', 'rb').read(),
      file_name="Hyetograph_Image.png",
      mime="image/jpeg",
      )
    
    btn_IDF_curves = st.download_button(
      label="Download IDF_curves Image",
      data=open('IDF_curves.png', 'rb').read(),
      file_name="IDF_curves_Image.png",
      mime="image/jpeg",
      )

                                                 
    
# Download image file


# Download pdf file



        
    
    
    